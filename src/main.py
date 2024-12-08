import os
import logging
from openpyxl import load_workbook
import pandas as pd
from config.config_loader import load_config, setup_logging
from utils.file_operations import create_export_folder, read_column_assignments
from processors.excel_processor import ExcelProcessor

def process_master_sheet(config):
    """Main function to process the master sheet."""
    try:
        # Load master sheet
        master_wb = load_workbook(config["master_file"])
        master_sheet = master_wb.active
        master_data = list(master_sheet.values)
        master_df = pd.DataFrame(master_data)

        # Validate master file
        if master_df.empty or master_df.isnull().all().all():
            raise ValueError("Master sheet is empty or invalid")

        # Initialize processor
        processor = ExcelProcessor(config, master_df)
        column_assignments = read_column_assignments(config["column_assignments"])

        # Process files in parallel
        with ThreadPoolExecutor() as executor:
            for file_name, mappings in column_assignments.items():
                executor.submit(processor.process_file, file_name, mappings)

        # Save the updated master file
        output_path = os.path.join(
            config["export_folder"],
            "2024--OUTPUT--Calculating Monthly Compensation Reporting.xlsx"
        )
        
        for row_idx, row_data in master_df.iterrows():
            for col_idx, value in enumerate(row_data, start=1):
                master_sheet.cell(row=row_idx + 1, column=col_idx, value=value)
        
        master_wb.save(output_path)
        logging.info(f"Master file saved to: {output_path}")

    except Exception as e:
        logging.error(f"Error processing master sheet: {e}")
        raise

def main():
    """Entry point of the application."""
    try:
        # Load configuration
        config = load_config()

        # Create export folder
        create_export_folder(config["export_folder"])

        # Setup logging
        setup_logging(config["export_folder"])

        # Process the master sheet
        process_master_sheet(config)

    except Exception as e:
        logging.error(f"Fatal error: {e}")
        raise

if __name__ == "__main__":
    main()